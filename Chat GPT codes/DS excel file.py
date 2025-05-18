from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# Define the output directory
output_dir = Path(r"C:\Users\Nike\Documents\Programming\Python\Chat GPT codes")  # Replace with your desired path
output_dir.mkdir(parents=True, exist_ok=True)  # Create the directory if it doesn't exist

# Create a new workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Data Science Roadmap"

# Define the header row
headers = ["Topic", "Sub-Topic", "Estimated Time", "Resources", "Completed"]
ws.append(headers)

# Apply bold font to the header row
header_font = Font(bold=True)
for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=5):
    for cell in col:
        cell.font = header_font

# Define the roadmap data
roadmap_data = [
    # Fundamentals
    ("Fundamentals", "", "3-4 months", "", ""),
    ("", "Mathematics for Data Science", "3-4 months", "Khan Academy, 3Blue1Brown", ""),
    ("", "", "Linear Algebra", "2 weeks", "Linear Algebra by 3Blue1Brown", ""),
    ("", "", "Statistics", "3 weeks", "Khan Academy, OpenIntro Statistics", ""),
    ("", "", "Calculus Basics", "1 week", "Khan Academy Calculus", ""),
    ("", "Programming Foundations", "4 weeks", "Automate the Boring Stuff with Python", ""),
    ("", "", "Python Basics", "4 weeks", "Python for Everybody (Coursera)", ""),
    ("", "", "Python Tools for Data Science", "", "", ""),
    ("", "", "", "NumPy", "1 week", "Official NumPy Documentation", ""),
    ("", "", "", "Pandas", "1 week", "Pandas by Wes McKinney", ""),
    ("", "", "", "Matplotlib & Seaborn", "2 weeks", "Python Data Science Handbook", ""),
    ("", "Introduction to Data Analysis", "2 weeks", "DataCamp", ""),
    ("", "", "Exploratory Data Analysis", "2 weeks", "Kaggle Tutorials", ""),
    ("", "", "Practice Basic Data Projects", "2 weeks", "Kaggle Datasets", ""),
    
    # Core Data Science
    ("Core Data Science", "", "4-6 months", "", ""),
    ("", "Machine Learning Basics", "6 weeks", "Andrew Ng's Machine Learning", ""),
    ("", "", "ML Concepts & Algorithms", "6 weeks", "Hands-On ML with Scikit-Learn", ""),
    ("", "", "Scikit-Learn", "2 weeks", "Scikit-Learn Documentation", ""),
    ("", "Data Visualization Techniques", "2 weeks", "Storytelling with Data", ""),
    ("", "", "Advanced Visualization Tools", "2 weeks", "Data Visualization Best Practices", ""),
    ("", "Advanced Python Tools", "", "", ""),
    ("", "", "SQL for Data Science", "3 weeks", "SQL for Data Science by UC Davis", ""),
    ("", "", "Regular Expressions", "1 week", "RegexOne", ""),
    
    # Specialization & Advanced Topics
    ("Specialization & Advanced Topics", "", "6-12 months", "", ""),
    ("", "Deep Learning", "6 months", "Deep Learning Specialization (Coursera)", ""),
    ("", "", "DL Frameworks (TensorFlow, PyTorch)", "6 months", "TensorFlow & PyTorch Docs", ""),
    ("", "", "Convolutional Neural Networks (CNN)", "6 months", "Deep Learning by Ian Goodfellow", ""),
    ("", "", "Natural Language Processing (NLP)", "6 months", "FastAI NLP Course", ""),
    ("", "Big Data Tools", "6 months", "Cloudera Hadoop", ""),
    ("", "", "Hadoop Ecosystem", "6 months", "Cloudera Hadoop", ""),
    ("", "", "Apache Spark", "6 months", "Spark: The Definitive Guide", ""),
    ("", "Domain Knowledge", "6 months", "Domain-specific Courses", ""),
]

# Add data to the sheet
for row in roadmap_data:
    ws.append(row)

# Add data validation for checkboxes in the 'Completed' column
completed_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws.add_data_validation(completed_validation)

# Apply validation to all rows in the 'Completed' column
for row in range(2, len(roadmap_data) + 2):
    completed_validation.add(ws[f"E{row}"])

# Apply conditional formatting: strike-through for completed tasks
strike_font = Font(strike=True)
dxf = DifferentialStyle(font=strike_font)
rule = Rule(type="expression", dxf=dxf, formula=["$E2=\"Yes\""])
ws.conditional_formatting.add(f"A2:E{len(roadmap_data) + 1}", rule)

# Save the enhanced Excel file
enhanced_excel_file = output_dir / "Enhanced_Data_Science_Roadmap.xlsx"
wb.save(enhanced_excel_file)

print(f"File saved at: {enhanced_excel_file}")
